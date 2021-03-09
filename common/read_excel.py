import os
import json
import openpyxl
from filelock import FileLock
from common.constant import Constant
from common.logger import logger

if not os.path.exists(Constant.EXCEL_DIR):
    os.mkdir(Constant.EXCEL_DIR)


class CasesData:

    def __init__(self, attrs):
        for t in attrs:
            try:
                setattr(self, t[0], t[1])
            except TypeError:
                continue

    def __str__(self):
        return getattr(self, "title", super().__str__())

    def __repr__(self):
        return getattr(self, "title", super().__repr__())

    def __setattr__(self, key, value):
        try:
            value = json.loads(value)
        except (json.decoder.JSONDecodeError, TypeError):
            try:
                if "[" in value and "]" in value:
                    value = eval(value)
            except (SyntaxError, TypeError):
                pass
        finally:
            super().__setattr__(key, value)


class ReadExcel:

    def __init__(self, file_path, *sheet_name):
        """

        :param file_path: 文件地址 -> str
        :param sheet_name: 工作表名字 -> str
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.wb = None
        self.sheet = dict()
        self.data = dict()

    def __open(self):
        # 创建操纵excel的实例
        with FileLock(self.file_path + ".lock", timeout=30):
            self.wb = openpyxl.load_workbook(self.file_path)
            for sheet in self.sheet_name:
                try:
                    self.sheet.setdefault(sheet, self.wb[sheet])
                except KeyError as e:
                    logger.error(f"The sheet: [{sheet}] does not exist. Track:{e}")
                    raise e
            self.wb.close()
        logger.info("Workbook: {}, Sheet: {}".format(self.file_path, self.sheet_name))

    def read_obj(self):
        self.__open()
        for k, y in self.sheet.items():
            data = self._read_excel_obj(y)
            self.data.setdefault(k, data)
        return self.data

    def read_dict(self):
        self.__open()
        data = dict()
        for k, y in self.sheet.items():
            dic = self._read_excel_dict(y)
            data.setdefault(k, dic)
        return data

    @staticmethod
    def _read_excel_dict(sheet) -> list:
        """
        方法一： 将每条数据以dict的形式储存于list中
        :param sheet:
        :return: 以列表形式存储的case_data -> list
        """
        # 获取首行数据，得到title并存于list
        title = []
        for i in sheet[1]:
            title.append(i.value)
        # 获取每一行的数据并与title组成dict
        case_data = []
        if sheet.max_row >= 3:
            for r in sheet[2:sheet.max_row]:
                case_list = []
                for va in r:
                    case_list.append(va.value)
                case_data.append(zip(title, case_list))
        else:
            data = [i.value for i in sheet[2:sheet.max_row]]
            case_data.append(zip(title, data))
        return case_data

    @staticmethod
    def _read_excel_obj(sheet_obj, *list_column) -> list:
        """
        自定义选择需要读取的column, 若未传则默认返回全部
        :param sheet_obj:
        :param list_column: 不定长参数， 需要读取的column(int) -> tuple
        :return: list中存着对象，对象中有实例属性 -> list
        """
        title = []
        if not list_column:  # 未传参时默认为获取全部
            for i in sheet_obj[1]:
                title.append(i.value)
            obj_data = []
            if sheet_obj.max_row >= 3:
                for r in sheet_obj[2:sheet_obj.max_row]:
                    case_list = []
                    for va in r:
                        case_list.append(va.value)
                    attr = CasesData(zip(title, case_list))
                    if getattr(attr, "id", None):
                        obj_data.append(attr)
                    else:
                        logger.warning(f"Invalid data exists in sheet:[{sheet_obj.title}],"
                                       f"The value of coordinate:[{r[0].coordinate}] is value:[{r[0].value}] ")
            else:
                data = [i.value for i in sheet_obj[2:sheet_obj.max_row]]
                attr = CasesData(zip(title, data))
                if getattr(attr, "id", None):
                    obj_data.append(attr)

        else:  # 指定column
            if sheet_obj.max_column < max(list_column):
                raise ValueError("Column out of range")
            data = list(sheet_obj.rows)
            for c in list_column:
                title.append(data[0][c - 1].value)
            obj_data = []
            for i in data[1:]:
                case_list = []
                for y in list_column:
                    x = y - 1
                    case_list.append(i[x].value)
                attr = CasesData(zip(title, case_list))
                obj_data.append(attr)
        return obj_data

    def write_data(self, sheet, *args) -> None:
        """
        写入excel
        :param sheet:
        :param args: 支持单个/多个单元格写入 Example: [{"row": int, "column": int, "msg": data}, ]
        :return:
        """

        def write(sheet, **kwargs):
            row = kwargs['row']
            column = kwargs['column']
            msg = kwargs['msg']
            try:
                logger.info("Excel write to data! row: {}, column: {}, value: {}".format(row, column, msg))
                sheet.cell(row=row, column=column, value=msg)
            except Exception as e:
                logger.error("Excel Write data function is error: {}".format(e))

        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb[sheet]
        for i in args:
            write(sheet, **i)
        wb.save(self.file_path)
        wb.close()


if __name__ == '__main__':
    # obj = ReadExcel(Constant.EXCEL_DIR + "/cases.xlsx", "smoke2", "smoke3", "smoke4", "smoke5", "smoke6", "smoke7",
    #                 "smoke8", "smoke9", "smoke10", "smoke11", "smoke12", "smoke13", "smoke14", "smoke15",
    #                 "smoke16", "smoke17")
    excel = ReadExcel(Constant.EXCEL_DIR + "/cases.xlsx", "init1")
    obj = excel.read_obj()
    dic = excel.read_dict()
    print(obj)
    print(list(dic))
